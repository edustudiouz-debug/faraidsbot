import os
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage

# Matplotlibni server muhitida (GUI'siz) xavfsiz ishlatish sozlamasi
import matplotlib
matplotlib.use('Agg')  
import matplotlib.pyplot as plt

from aiogram import Router, F, types
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import FSInputFile, ReplyKeyboardRemove

from config import ADMINS
from database import connect_db
from handlers.common import admin_keyboard
from utils.pdf_generator import create_medical_pdf  # PDF generator moduli ulandi

router = Router()

# FSM (State) Holatlari
class AdminStates(StatesGroup):
    waiting_for_broadcast = State()
    waiting_for_ban_id = State()
    
    # PDF yaratish uchun holatlar
    waiting_pdf_user_id = State()
    waiting_pdf_name = State()
    waiting_pdf_analysis_type = State()
    waiting_pdf_result = State()

# =====================================================================
# 1. 📊 FOIZLI VA GRAFIKLI MUKAMMAL STATISTIKA (EXCEL + RASM)
# =====================================================================
@router.message(F.text == "📊 Statistika", F.from_user.id.in_(ADMINS))
async def admin_stats(message: types.Message):
    await message.answer("📊 Bazadan ma'lumotlar yig'ilmoqda va grafik chizilmoqda, iltimos kuting...")
    
    conn = connect_db()
    total_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    
    if total_users == 0:
        conn.close()
        await message.answer("⚠️ Botda hali foydalanuvchilar mavjud emas.", reply_markup=admin_keyboard)
        return
        
    males = conn.execute("SELECT COUNT(*) FROM users WHERE gender='Erkak'").fetchone()[0]
    females = conn.execute("SELECT COUNT(*) FROM users WHERE gender='Ayol'").fetchone()[0]
    
    seven_days_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
    new_users = conn.execute(
        "SELECT COUNT(*) FROM users WHERE user_id NOT IN (SELECT user_id FROM analysis_requests WHERE date_requested < ?)", 
        (seven_days_ago,)
    ).fetchone()[0]
    
    analysis_users = conn.execute("SELECT COUNT(DISTINCT user_id) FROM analysis_requests").fetchone()[0]
    appointed_users = conn.execute("SELECT COUNT(DISTINCT user_id) FROM appointments").fetchone()[0]
    conn.close()
    
    male_pct = (males / total_users) * 100 if total_users > 0 else 0
    female_pct = (females / total_users) * 100 if total_users > 0 else 0
    new_users_pct = (new_users / total_users) * 100 if total_users > 0 else 0
    analysis_pct = (analysis_users / total_users) * 100 if total_users > 0 else 0
    appointed_pct = (appointed_users / total_users) * 100 if total_users > 0 else 0
    
    graph_filename = "statistika_grafik.png"
    labels = ['Erkaklar', 'Ayollar', 'Yangi a\'zolar', 'Tahlil so\'raganlar', 'Qabulga yozilganlar']
    sizes = [male_pct, female_pct, new_users_pct, analysis_pct, appointed_pct]
    colors = ['#3498db', '#e74c3c', '#2ecc71', '#f1c40f', '#9b59b6']
    
    plt.figure(figsize=(7, 5))
    plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140, colors=colors, shadow=True)
    plt.title(f"Farg'ona OITS Markazi Boti Statistikasi\n(Jami a'zolar: {total_users} ta)", fontsize=13, fontweight='bold')
    plt.axis('equal')
    plt.tight_layout()
    plt.savefig(graph_filename, dpi=100)
    plt.close()
    
    excel_filename = "Markaz_Bot_Statistika.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Umumiy Statistika"
    
    ws.merge_cells('A1:C1')
    ws['A1'] = "FARG'ONA VILOYAT OITS MARKAZI BOTI STATISTIK HISOBOTI"
    ws['A1'].font = openpyxl.styles.Font(size=12, bold=True, color="FFFFFF")
    ws['A1'].fill = openpyxl.styles.PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal="center")
    
    headers = ["Ko'rsatkich nomi", "Soni (ta)", "Ulushi (Foizda %)"]
    ws.append([]) 
    ws.append(headers)
    
    for col in range(1, 4):
        cell = ws.cell(row=3, column=col)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
    
    data_rows = [
        ["Jami Bot Foydalanuvchilari", total_users, "100%"],
        ["👨 Erkaklar soni", males, f"{male_pct:.1f}%"],
        ["👩 Ayollar soni", females, f"{female_pct:.1f}%"],
        ["🆕 Yangi foydalanuvchilar (7 kunlik)", new_users, f"{new_users_pct:.1f}%"],
        ["🧾 Tahlil natijasini so'raganlar", analysis_users, f"{analysis_pct:.1f}%"],
        ["📅 Shifokor qabuliga yozilganlar", appointed_users, f"{appointed_pct:.1f}%"]
    ]
    
    for row in data_rows:
        ws.append(row)
        
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    
    img = OpenpyxlImage(graph_filename)
    ws.add_image(img, 'E3')
    wb.save(excel_filename)
    
    caption_text = (
        f"📊 <b>Kengaytirilgan Statistika Hisoboti:</b>\n\n"
        f"👥 Jami a'zolar: <b>{total_users} ta</b> (100%)\n"
        f"👨 Erkaklar: {males} ta (<b>{male_pct:.1f}%</b>)\n"
        f"👩 Ayollar: {females} ta (<b>{female_pct:.1f}%</b>)\n"
        f"🆕 Yangi a'zolar: {new_users} ta (<b>{new_users_pct:.1f}%</b>)\n"
        f"🧾 Tahlil olganlar: {analysis_users} ta (<b>{analysis_pct:.1f}%</b>)\n"
        f"📅 Qabulga yozilganlar: {appointed_users} ta (<b>{appointed_pct:.1f}%</b>)\n\n"
        f"📥 Batafsil grafik diagramma va raqamlar pastdagi Excel fayl ichiga joylandi!"
    )
    
    await message.bot.send_photo(chat_id=message.from_user.id, photo=FSInputFile(graph_filename), caption=caption_text)
    await message.bot.send_document(chat_id=message.from_user.id, document=FSInputFile(excel_filename), reply_markup=admin_keyboard)
    
    if os.path.exists(graph_filename): os.remove(graph_filename)
    if os.path.exists(excel_filename): os.remove(excel_filename)

# =====================================================================
# 2. 📢 XABAR YUBORISH TIZIMI (RASSILKA)
# =====================================================================
@router.message(F.text == "📢 Xabar yuborish", F.from_user.id.in_(ADMINS))
async def admin_bc_start(message: types.Message, state: FSMContext):
    await state.set_state(AdminStates.waiting_for_broadcast)
    await message.answer("📝 Barcha foydalanuvchilarga yuboriladigan xabar matnini kiriting:", reply_markup=ReplyKeyboardRemove())

@router.message(AdminStates.waiting_for_broadcast, F.from_user.id.in_(ADMINS))
async def process_bc(message: types.Message, state: FSMContext):
    await state.clear()
    conn = connect_db()
    users = conn.execute("SELECT user_id FROM users WHERE is_banned=0").fetchall()
    conn.close()
    
    count = 0
    for r in users:
        try: 
            await message.bot.send_message(chat_id=r[0], text=message.text)
            count += 1
        except Exception: 
            continue
            
    await message.answer(f"✅ Xabar {count} ta faol foydalanuvchiga muvaffaqiyatli yuborildi.", reply_markup=admin_keyboard)

# =====================================================================
# 3. 🚫 BLOKLASH TIZIMI (BAN)
# =====================================================================
@router.message(F.text == "🚫 Bloklash (Ban)", F.from_user.id.in_(ADMINS))
async def admin_ban_start(message: types.Message, state: FSMContext):
    await state.set_state(AdminStates.waiting_for_ban_id)
    await message.answer("🔒 Bloklamoqchi bo'lgan foydalanuvchining <b>Telegram ID</b> raqamini kiriting:", reply_markup=ReplyKeyboardRemove())

@router.message(AdminStates.waiting_for_ban_id, F.from_user.id.in_(ADMINS))
async def process_ban(message: types.Message, state: FSMContext):
    await state.clear()
    user_id_input = message.text.strip()
    
    if not user_id_input.isdigit():
        await message.answer("⚠️ ID raqami faqat raqamlardan iborat bo'lishi kerak!", reply_markup=admin_keyboard)
        return
        
    conn = connect_db()
    conn.execute("INSERT OR REPLACE INTO users (user_id, is_banned) VALUES (?, 1)", (int(user_id_input),))
    conn.commit()
    conn.close()
    
    await message.answer(f"🚫 Foydalanuvchi (ID: {user_id_input}) muvaffaqiyatli bloklandi.", reply_markup=admin_keyboard)

# =====================================================================
# 4. 📋 BEMORLAR NAVBATI (BAZA TUZILISHIGA MOSLASHGAN)
# =====================================================================
@router.message(F.text == "📋 Bemorlar navbati", F.from_user.id.in_(ADMINS))
async def admin_view_queue(message: types.Message):
    conn = connect_db()
    try:
        # Barcha ustunlarni olish
        cursor = conn.execute("SELECT id, user_id, full_name, doctor_type, details, phone_number, date_booked FROM appointments")
        appointments = cursor.fetchall()
        
        if not appointments:
            await message.answer("📋 Hozirda navbatda bemorlar yo'q.", reply_markup=admin_keyboard)
            conn.close()
            return

        text = "📋 <b>SHIFOKOR QABULIDAGI BEMORLAR RO'YXATI:</b>\n\n"
        
        for idx, app in enumerate(appointments, 1):
            # Indekslar: 
            # app[0]=ID, app[1]=user_id, app[2]=full_name, 
            # app[3]=doctor_type, app[4]=details(vaqt/matn), app[5]=phone, app[6]=date
            
            # Matnlarni tozalash (agar kerak bo'lsa)
            details_clean = str(app[4]).replace("Bemor:", "").replace("Vaqt:", "").strip()
            
            text += (
                f"<b>{idx}. BEMOR: {app[2]}</b>\n"      # full_name
                f"👨‍⚕️ Shifokor: {app[3]}\n"             # doctor_type
                f"📅 Vaqt: {details_clean}\n"   # details
                f"📞 Tel: {app[5]}\n"                    # phone_number
                f"🆔 Telegram ID: <code>{app[1]}</code>\n" # user_id
                f"🕒 Yozilgan sana: {app[6]}\n"          # date_booked
                f"_________________________________\n"
            )
        
        await message.answer(text, reply_markup=admin_keyboard, parse_mode="HTML")
        
    except Exception as e:
        await message.answer(f"❌ Xatolik: {e}")
    finally:
        conn.close()
# =====================================================================
# 5. 📊 TAHLIL SO'RAGANLAR (EXCEL - YANGI QO'SHILDI)
# =====================================================================
@router.message(F.text == "📊 Tahlil so'raganlar (Excel)", F.from_user.id.in_(ADMINS))
async def admin_get_excel_report(message: types.Message):
    await message.answer("📊 Shifrlangan tahlil so'rovlari Excelga yuklanmoqda...")
    
    conn = connect_db()
    requests = conn.execute("SELECT id, user_id, passport, jshshir, date_requested FROM analysis_requests").fetchall()
    conn.close()
    
    if not requests:
        await message.answer("⚠️ Tahlil so'rovlari ma'lumotlar bazasida mavjud emas.", reply_markup=admin_keyboard)
        return
        
    # Bazadagi ma'lumotlarni dekodlash funksiyasi (Agar shifrlangan bo'lsa shifrdan ochadi)
    from database import decrypt_data  # database ichida decrypt_data borligiga ishonch hosil qiling
    
    clean_data = []
    for r in requests:
        try:
            pass_dec = decrypt_data(r[2])
            jsh_dec = decrypt_data(r[3])
        except Exception:
            pass_dec, jsh_dec = r[2], r[3]  # Agar shifrlanmagan bo'lsa boricha oladi
            
        clean_data.append({
            "So'rov ID": r[0],
            "Foydalanuvchi ID": r[1],
            "Pasport Seriya": pass_dec,
            "JSHSHIR Kod": jsh_dec,
            "Sana": r[4]
        })
        
    df = pd.DataFrame(clean_data)
    report_file = "Tahlil_So_raganlar_Hisoboti.xlsx"
    df.to_excel(report_file, index=False)
    
    await message.bot.send_document(
        chat_id=message.from_user.id, 
        document=FSInputFile(report_file), 
        caption="✅ Barcha tahlil so'ragan bemorlar ro'yxati (Pasport va JSHSHIR shifrdan ochilgan holda).",
        reply_markup=admin_keyboard
    )
    if os.path.exists(report_file): os.remove(report_file)

# =====================================================================
# 6. 📄 PDF TAHLIL YARATISH (Avtomatik generatsiya)
# =====================================================================
@router.message(F.text == "📄 PDF Tahlil yaratish", F.from_user.id.in_(ADMINS))
async def admin_pdf_start(message: types.Message, state: FSMContext):
    await state.set_state(AdminStates.waiting_pdf_user_id)
    await message.answer("📄 Bemorning Telegram ID raqamini kiriting:", reply_markup=ReplyKeyboardRemove())

# ... (waiting_pdf_user_id, name, type, result holatlari xuddi o'z holicha qoladi) ...

@router.message(AdminStates.waiting_pdf_result, F.from_user.id.in_(ADMINS))
async def admin_pdf_final(message: types.Message, state: FSMContext):
    data = await state.get_data()
    await state.clear()
    
    user_id = data['pdf_user_id']
    name = data['pdf_name']
    analysis_type = data['pdf_type']
    result = message.text
    
    await message.answer("⏳ PDF hujjat generatsiya qilinmoqda...")
    pdf_path = create_medical_pdf(name, analysis_type, result)
    
    if pdf_path and os.path.exists(pdf_path):
        await message.bot.send_document(message.from_user.id, FSInputFile(pdf_path), caption="✅ Tayyor PDF:", reply_markup=admin_keyboard)
        try:
            await message.bot.send_document(user_id, FSInputFile(pdf_path), caption="🧾 Sizning rasmiy tahlil natijangiz.")
        except:
            await message.answer("⚠️ Bemorga yuborilmadi.")
        os.remove(pdf_path)

# =====================================================================
# 7. 🧾 TEST TAHLILI YUKLASH (Alohida fayl yuborish tizimi)
# =====================================================================
# Bu qism endi "Test tahlili yuklash" uchun alohida ishlaydi
class AdminUploadStates(StatesGroup):
    waiting_id = State()
    waiting_file = State()

@router.message(F.text == "🧾 Test tahlili yuklash", F.from_user.id.in_(ADMINS))
async def admin_upload_start(message: types.Message, state: FSMContext):
    await state.set_state(AdminUploadStates.waiting_id)
    await message.answer("👤 Fayl yuborilishi kerak bo'lgan bemorning ID raqamini kiriting:")

@router.message(AdminUploadStates.waiting_id, F.from_user.id.in_(ADMINS))
async def upload_id(message: types.Message, state: FSMContext):
    await state.update_data(target_id=message.text)
    await state.set_state(AdminUploadStates.waiting_file)
    await message.answer("📎 Endi tayyor tahlil faylini (PDF yoki Rasm) yuboring:")

@router.message(AdminUploadStates.waiting_file, F.from_user.id.in_(ADMINS))
async def upload_file_final(message: types.Message, state: FSMContext):
    data = await state.get_data()
    target_id = data['target_id']
    await state.clear()
    
    try:
        if message.document:
            await message.bot.send_document(target_id, message.document.file_id, caption="🧾 Tayyor tahlil natijasi.")
        elif message.photo:
            await message.bot.send_photo(target_id, message.photo[-1].file_id, caption="🧾 Tayyor tahlil natijasi.")
        else:
            await message.answer("❌ Iltimos, faqat fayl yoki rasm yuboring!")
            return
        await message.answer("✅ Fayl muvaffaqiyatli yuborildi!", reply_markup=admin_keyboard)
    except Exception as e:
        await message.answer(f"❌ Xatolik: {e}")
        
# =====================================================================
# 8. 🔔 TAHLIL SO'ROVLARINI ADMINGA YUBORISH VA TASDIQLASH
# =====================================================================

# Adminga so'rovni tugmalar bilan yuborish funksiyasi
async def send_new_analysis_request_to_admin(bot, user_id, name, passport, jshshir, birth_date):
    from config import ADMINS
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Qabul qilish", callback_data=f"accept_{user_id}"),
            InlineKeyboardButton(text="❌ Rad etish", callback_data=f"reject_{user_id}")
        ]
    ])
    
    text = (
        f"🔔 <b>YANGI TAHLIL SO'ROVI (FACE CONTROL)</b>\n\n"
        f"👤 Bemor: {name}\n"
        f"🆔 Telegram ID: {user_id}\n"
        f"🪪 Pasport: {passport}\n"
        f"🔢 JSHSHIR: {jshshir}\n"
        f"📅 Tug'ilgan sana: {birth_date}"
    )
    
    for admin_id in ADMINS:
        try:
            await bot.send_message(chat_id=admin_id, text=text, reply_markup=keyboard, parse_mode="HTML")
        except:
            continue

# Admin bosgan tugmalarni qayta ishlash
import os
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage

# Matplotlibni server muhitida (GUI'siz) xavfsiz ishlatish sozlamasi
import matplotlib
matplotlib.use('Agg')  
import matplotlib.pyplot as plt

from aiogram import Router, F, types
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import FSInputFile, ReplyKeyboardRemove

from config import ADMINS
from database import connect_db
from handlers.common import admin_keyboard
from utils.pdf_generator import create_medical_pdf  # PDF generator moduli ulandi

router = Router()

# FSM (State) Holatlari
class AdminStates(StatesGroup):
    waiting_for_broadcast = State()
    waiting_for_ban_id = State()
    
    # PDF yaratish uchun holatlar
    waiting_pdf_user_id = State()
    waiting_pdf_name = State()
    waiting_pdf_analysis_type = State()
    waiting_pdf_result = State()

# =====================================================================
# 1. 📊 FOIZLI VA GRAFIKLI MUKAMMAL STATISTIKA (EXCEL + RASM)
# =====================================================================
@router.message(F.text == "📊 Statistika", F.from_user.id.in_(ADMINS))
async def admin_stats(message: types.Message):
    await message.answer("📊 Bazadan ma'lumotlar yig'ilmoqda va grafik chizilmoqda, iltimos kuting...")
    
    conn = connect_db()
    total_users = conn.execute("SELECT COUNT(*) FROM users").fetchone()[0]
    
    if total_users == 0:
        conn.close()
        await message.answer("⚠️ Botda hali foydalanuvchilar mavjud emas.", reply_markup=admin_keyboard)
        return
        
    males = conn.execute("SELECT COUNT(*) FROM users WHERE gender='Erkak'").fetchone()[0]
    females = conn.execute("SELECT COUNT(*) FROM users WHERE gender='Ayol'").fetchone()[0]
    
    seven_days_ago = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d %H:%M:%S')
    new_users = conn.execute(
        "SELECT COUNT(*) FROM users WHERE user_id NOT IN (SELECT user_id FROM analysis_requests WHERE date_requested < ?)", 
        (seven_days_ago,)
    ).fetchone()[0]
    
    analysis_users = conn.execute("SELECT COUNT(DISTINCT user_id) FROM analysis_requests").fetchone()[0]
    appointed_users = conn.execute("SELECT COUNT(DISTINCT user_id) FROM appointments").fetchone()[0]
    conn.close()
    
    male_pct = (males / total_users) * 100 if total_users > 0 else 0
    female_pct = (females / total_users) * 100 if total_users > 0 else 0
    new_users_pct = (new_users / total_users) * 100 if total_users > 0 else 0
    analysis_pct = (analysis_users / total_users) * 100 if total_users > 0 else 0
    appointed_pct = (appointed_users / total_users) * 100 if total_users > 0 else 0
    
    graph_filename = "statistika_grafik.png"
    labels = ['Erkaklar', 'Ayollar', 'Yangi a\'zolar', 'Tahlil so\'raganlar', 'Qabulga yozilganlar']
    sizes = [male_pct, female_pct, new_users_pct, analysis_pct, appointed_pct]
    colors = ['#3498db', '#e74c3c', '#2ecc71', '#f1c40f', '#9b59b6']
    
    plt.figure(figsize=(7, 5))
    plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140, colors=colors, shadow=True)
    plt.title(f"Farg'ona OITS Markazi Boti Statistikasi\n(Jami a'zolar: {total_users} ta)", fontsize=13, fontweight='bold')
    plt.axis('equal')
    plt.tight_layout()
    plt.savefig(graph_filename, dpi=100)
    plt.close()
    
    excel_filename = "Markaz_Bot_Statistika.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Umumiy Statistika"
    
    ws.merge_cells('A1:C1')
    ws['A1'] = "FARG'ONA VILOYAT OITS MARKAZI BOTI STATISTIK HISOBOTI"
    ws['A1'].font = openpyxl.styles.Font(size=12, bold=True, color="FFFFFF")
    ws['A1'].fill = openpyxl.styles.PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    ws['A1'].alignment = openpyxl.styles.Alignment(horizontal="center")
    
    headers = ["Ko'rsatkich nomi", "Soni (ta)", "Ulushi (Foizda %)"]
    ws.append([]) 
    ws.append(headers)
    
    for col in range(1, 4):
        cell = ws.cell(row=3, column=col)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
    
    data_rows = [
        ["Jami Bot Foydalanuvchilari", total_users, "100%"],
        ["👨 Erkaklar soni", males, f"{male_pct:.1f}%"],
        ["👩 Ayollar soni", females, f"{female_pct:.1f}%"],
        ["🆕 Yangi foydalanuvchilar (7 kunlik)", new_users, f"{new_users_pct:.1f}%"],
        ["🧾 Tahlil natijasini so'raganlar", analysis_users, f"{analysis_pct:.1f}%"],
        ["📅 Shifokor qabuliga yozilganlar", appointed_users, f"{appointed_pct:.1f}%"]
    ]
    
    for row in data_rows:
        ws.append(row)
        
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    
    img = OpenpyxlImage(graph_filename)
    ws.add_image(img, 'E3')
    wb.save(excel_filename)
    
    caption_text = (
        f"📊 <b>Kengaytirilgan Statistika Hisoboti:</b>\n\n"
        f"👥 Jami a'zolar: <b>{total_users} ta</b> (100%)\n"
        f"👨 Erkaklar: {males} ta (<b>{male_pct:.1f}%</b>)\n"
        f"👩 Ayollar: {females} ta (<b>{female_pct:.1f}%</b>)\n"
        f"🆕 Yangi a'zolar: {new_users} ta (<b>{new_users_pct:.1f}%</b>)\n"
        f"🧾 Tahlil olganlar: {analysis_users} ta (<b>{analysis_pct:.1f}%</b>)\n"
        f"📅 Qabulga yozilganlar: {appointed_users} ta (<b>{appointed_pct:.1f}%</b>)\n\n"
        f"📥 Batafsil grafik diagramma va raqamlar pastdagi Excel fayl ichiga joylandi!"
    )
    
    await message.bot.send_photo(chat_id=message.from_user.id, photo=FSInputFile(graph_filename), caption=caption_text)
    await message.bot.send_document(chat_id=message.from_user.id, document=FSInputFile(excel_filename), reply_markup=admin_keyboard)
    
    if os.path.exists(graph_filename): os.remove(graph_filename)
    if os.path.exists(excel_filename): os.remove(excel_filename)

# =====================================================================
# 2. 📢 XABAR YUBORISH TIZIMI (RASSILKA)
# =====================================================================
@router.message(F.text == "📢 Xabar yuborish", F.from_user.id.in_(ADMINS))
async def admin_bc_start(message: types.Message, state: FSMContext):
    await state.set_state(AdminStates.waiting_for_broadcast)
    await message.answer("📝 Barcha foydalanuvchilarga yuboriladigan xabar matnini kiriting:", reply_markup=ReplyKeyboardRemove())

@router.message(AdminStates.waiting_for_broadcast, F.from_user.id.in_(ADMINS))
async def process_bc(message: types.Message, state: FSMContext):
    await state.clear()
    conn = connect_db()
    users = conn.execute("SELECT user_id FROM users WHERE is_banned=0").fetchall()
    conn.close()
    
    count = 0
    for r in users:
        try: 
            await message.bot.send_message(chat_id=r[0], text=message.text)
            count += 1
        except Exception: 
            continue
            
    await message.answer(f"✅ Xabar {count} ta faol foydalanuvchiga muvaffaqiyatli yuborildi.", reply_markup=admin_keyboard)

# =====================================================================
# 3. 🚫 BLOKLASH TIZIMI (BAN)
# =====================================================================
@router.message(F.text == "🚫 Bloklash (Ban)", F.from_user.id.in_(ADMINS))
async def admin_ban_start(message: types.Message, state: FSMContext):
    await state.set_state(AdminStates.waiting_for_ban_id)
    await message.answer("🔒 Bloklamoqchi bo'lgan foydalanuvchining <b>Telegram ID</b> raqamini kiriting:", reply_markup=ReplyKeyboardRemove())

@router.message(AdminStates.waiting_for_ban_id, F.from_user.id.in_(ADMINS))
async def process_ban(message: types.Message, state: FSMContext):
    await state.clear()
    user_id_input = message.text.strip()
    
    if not user_id_input.isdigit():
        await message.answer("⚠️ ID raqami faqat raqamlardan iborat bo'lishi kerak!", reply_markup=admin_keyboard)
        return
        
    conn = connect_db()
    conn.execute("INSERT OR REPLACE INTO users (user_id, is_banned) VALUES (?, 1)", (int(user_id_input),))
    conn.commit()
    conn.close()
    
    await message.answer(f"🚫 Foydalanuvchi (ID: {user_id_input}) muvaffaqiyatli bloklandi.", reply_markup=admin_keyboard)

# =====================================================================
# 4. 📋 BEMORLAR NAVBATI (BAZA TUZILISHIGA MOSLASHGAN)
# =====================================================================
@router.message(F.text == "📋 Bemorlar navbati", F.from_user.id.in_(ADMINS))
async def admin_view_queue(message: types.Message):
    conn = connect_db()
    try:
        # Barcha ustunlarni olish
        cursor = conn.execute("SELECT id, user_id, full_name, doctor_type, details, phone_number, date_booked FROM appointments")
        appointments = cursor.fetchall()
        
        if not appointments:
            await message.answer("📋 Hozirda navbatda bemorlar yo'q.", reply_markup=admin_keyboard)
            conn.close()
            return

        text = "📋 <b>SHIFOKOR QABULIDAGI BEMORLAR RO'YXATI:</b>\n\n"
        
        for idx, app in enumerate(appointments, 1):
            # Indekslar: 
            # app[0]=ID, app[1]=user_id, app[2]=full_name, 
            # app[3]=doctor_type, app[4]=details(vaqt/matn), app[5]=phone, app[6]=date
            
            # Matnlarni tozalash (agar kerak bo'lsa)
            details_clean = str(app[4]).replace("Bemor:", "").replace("Vaqt:", "").strip()
            
            text += (
                f"<b>{idx}. BEMOR: {app[2]}</b>\n"      # full_name
                f"👨‍⚕️ Shifokor: {app[3]}\n"             # doctor_type
                f"📅 Vaqt: {details_clean}\n"   # details
                f"📞 Tel: {app[5]}\n"                    # phone_number
                f"🆔 Telegram ID: <code>{app[1]}</code>\n" # user_id
                f"🕒 Yozilgan sana: {app[6]}\n"          # date_booked
                f"_________________________________\n"
            )
        
        await message.answer(text, reply_markup=admin_keyboard, parse_mode="HTML")
        
    except Exception as e:
        await message.answer(f"❌ Xatolik: {e}")
    finally:
        conn.close()
# =====================================================================
# 5. 📊 TAHLIL SO'RAGANLAR (EXCEL - YANGI QO'SHILDI)
# =====================================================================
@router.message(F.text == "📊 Tahlil so'raganlar (Excel)", F.from_user.id.in_(ADMINS))
async def admin_get_excel_report(message: types.Message):
    await message.answer("📊 Shifrlangan tahlil so'rovlari Excelga yuklanmoqda...")
    
    conn = connect_db()
    requests = conn.execute("SELECT id, user_id, passport, jshshir, date_requested FROM analysis_requests").fetchall()
    conn.close()
    
    if not requests:
        await message.answer("⚠️ Tahlil so'rovlari ma'lumotlar bazasida mavjud emas.", reply_markup=admin_keyboard)
        return
        
    # Bazadagi ma'lumotlarni dekodlash funksiyasi (Agar shifrlangan bo'lsa shifrdan ochadi)
    from database import decrypt_data  # database ichida decrypt_data borligiga ishonch hosil qiling
    
    clean_data = []
    for r in requests:
        try:
            pass_dec = decrypt_data(r[2])
            jsh_dec = decrypt_data(r[3])
        except Exception:
            pass_dec, jsh_dec = r[2], r[3]  # Agar shifrlanmagan bo'lsa boricha oladi
            
        clean_data.append({
            "So'rov ID": r[0],
            "Foydalanuvchi ID": r[1],
            "Pasport Seriya": pass_dec,
            "JSHSHIR Kod": jsh_dec,
            "Sana": r[4]
        })
        
    df = pd.DataFrame(clean_data)
    report_file = "Tahlil_So_raganlar_Hisoboti.xlsx"
    df.to_excel(report_file, index=False)
    
    await message.bot.send_document(
        chat_id=message.from_user.id, 
        document=FSInputFile(report_file), 
        caption="✅ Barcha tahlil so'ragan bemorlar ro'yxati (Pasport va JSHSHIR shifrdan ochilgan holda).",
        reply_markup=admin_keyboard
    )
    if os.path.exists(report_file): os.remove(report_file)

# =====================================================================
# 6. 📄 PDF TAHLIL YARATISH (Avtomatik generatsiya)
# =====================================================================
@router.message(F.text == "📄 PDF Tahlil yaratish", F.from_user.id.in_(ADMINS))
async def admin_pdf_start(message: types.Message, state: FSMContext):
    await state.set_state(AdminStates.waiting_pdf_user_id)
    await message.answer("📄 Bemorning Telegram ID raqamini kiriting:", reply_markup=ReplyKeyboardRemove())

# ... (waiting_pdf_user_id, name, type, result holatlari xuddi o'z holicha qoladi) ...

@router.message(AdminStates.waiting_pdf_result, F.from_user.id.in_(ADMINS))
async def admin_pdf_final(message: types.Message, state: FSMContext):
    data = await state.get_data()
    await state.clear()
    
    user_id = data['pdf_user_id']
    name = data['pdf_name']
    analysis_type = data['pdf_type']
    result = message.text
    
    await message.answer("⏳ PDF hujjat generatsiya qilinmoqda...")
    pdf_path = create_medical_pdf(name, analysis_type, result)
    
    if pdf_path and os.path.exists(pdf_path):
        await message.bot.send_document(message.from_user.id, FSInputFile(pdf_path), caption="✅ Tayyor PDF:", reply_markup=admin_keyboard)
        try:
            await message.bot.send_document(user_id, FSInputFile(pdf_path), caption="🧾 Sizning rasmiy tahlil natijangiz.")
        except:
            await message.answer("⚠️ Bemorga yuborilmadi.")
        os.remove(pdf_path)

# =====================================================================
# 7. 🧾 TEST TAHLILI YUKLASH (Alohida fayl yuborish tizimi)
# =====================================================================
# Bu qism endi "Test tahlili yuklash" uchun alohida ishlaydi
class AdminUploadStates(StatesGroup):
    waiting_id = State()
    waiting_file = State()

@router.message(F.text == "🧾 Test tahlili yuklash", F.from_user.id.in_(ADMINS))
async def admin_upload_start(message: types.Message, state: FSMContext):
    await state.set_state(AdminUploadStates.waiting_id)
    await message.answer("👤 Fayl yuborilishi kerak bo'lgan bemorning ID raqamini kiriting:")

@router.message(AdminUploadStates.waiting_id, F.from_user.id.in_(ADMINS))
async def upload_id(message: types.Message, state: FSMContext):
    await state.update_data(target_id=message.text)
    await state.set_state(AdminUploadStates.waiting_file)
    await message.answer("📎 Endi tayyor tahlil faylini (PDF yoki Rasm) yuboring:")

@router.message(AdminUploadStates.waiting_file, F.from_user.id.in_(ADMINS))
async def upload_file_final(message: types.Message, state: FSMContext):
    data = await state.get_data()
    target_id = data['target_id']
    await state.clear()
    
    try:
        if message.document:
            await message.bot.send_document(target_id, message.document.file_id, caption="🧾 Tayyor tahlil natijasi.")
        elif message.photo:
            await message.bot.send_photo(target_id, message.photo[-1].file_id, caption="🧾 Tayyor tahlil natijasi.")
        else:
            await message.answer("❌ Iltimos, faqat fayl yoki rasm yuboring!")
            return
        await message.answer("✅ Fayl muvaffaqiyatli yuborildi!", reply_markup=admin_keyboard)
    except Exception as e:
        await message.answer(f"❌ Xatolik: {e}")
        
# =====================================================================
# 8. 🔔 TAHLIL SO'ROVLARINI ADMINGA YUBORISH VA TASDIQLASH
# =====================================================================

# Adminga so'rovni tugmalar bilan yuborish funksiyasi
async def send_new_analysis_request_to_admin(bot, user_id, name, passport, jshshir, birth_date):
    from config import ADMINS
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="✅ Qabul qilish", callback_data=f"accept_{user_id}"),
            InlineKeyboardButton(text="❌ Rad etish", callback_data=f"reject_{user_id}")
        ]
    ])
    
    text = (
        f"🔔 <b>YANGI TAHLIL SO'ROVI (FACE CONTROL)</b>\n\n"
        f"👤 Bemor: {name}\n"
        f"🆔 Telegram ID: {user_id}\n"
        f"🪪 Pasport: {passport}\n"
        f"🔢 JSHSHIR: {jshshir}\n"
        f"📅 Tug'ilgan sana: {birth_date}"
    )
    
    for admin_id in ADMINS:
        try:
            await bot.send_message(chat_id=admin_id, text=text, reply_markup=keyboard, parse_mode="HTML")
        except:
            continue

# Admin bosgan tugmalarni qayta ishlash
@router.callback_query(F.data.startswith(("accept_", "reject_")))
async def process_admin_decision(callback: types.CallbackQuery):
    action, user_id = callback.data.split("_")
    user_id = int(user_id)
    
    # HOLATNI TANLASH
    status_text = "✅ <b>Holat: Qabul qilindi.</b>" if action == "accept" else "❌ <b>Holat: Rad etildi.</b>"

    # 1. Rasm yoki matnligiga qarab tahrirlash
    try:
        if callback.message.photo:
            # Rasm bo'lsa: caption (sarlavha) ni yangilaymiz
            new_caption = (callback.message.caption or "") + f"\n\n{status_text}"
            await callback.message.edit_caption(caption=new_caption, parse_mode="HTML")
        else:
            # Matn bo'lsa: text (matn) ni yangilaymiz
            new_text = (callback.message.text or "") + f"\n\n{status_text}"
            await callback.message.edit_text(text=new_text, parse_mode="HTML")
    except Exception as e:
        print(f"Tahrirlashda xatolik: {e}")

    # 2. Foydalanuvchiga xabar yuborish
    if action == "accept":
        msg = "✅ <b>Face ID tasdiqlandi!</b>\n\nSiz tizimdan muvaffaqiyatli o'tdingiz. Birozdan so'ng taxlil natijasi paydo bo'ladi !"
    else:
        msg = "❌ Afsuski, Face ID so'rovingiz rad etildi. Iltimos, ma'lumotlaringizni tekshiring."

    try:
        await callback.bot.send_message(chat_id=user_id, text=msg, parse_mode="HTML")
    except Exception as e:
        await callback.answer(f"⚠️ Foydalanuvchiga xabar bormadi: {e}", show_alert=True)
            
    await callback.answer()